import os
from typing import List, Optional
from fastapi import FastAPI, Header, HTTPException, UploadFile, File
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy import create_engine, Column, Integer, String, DateTime
from sqlalchemy.orm import declarative_base, sessionmaker
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

DATABASE_URL = os.getenv("DATABASE_URL", "")
IMPORT_TOKEN = os.getenv("IMPORT_TOKEN", "")

if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()

class Job(Base):
    __tablename__ = "jobs"
    id = Column(Integer, primary_key=True)
    job_number = Column(String, unique=True, index=True)
    job_name = Column(String)
    customer = Column(String)
    pm = Column(String)
    address = Column(String)
    city = Column(String)
    state_zip = Column(String)
    start_date = Column(String)
    date_entered = Column(String)
    duration_days = Column(String)
    laborers_needed = Column(String)
    operators_needed = Column(String)
    job_type = Column(String)
    priority = Column(String)
    status = Column(String, default="Active")
    updated_at = Column(DateTime, default=datetime.utcnow)

class Employee(Base):
    __tablename__ = "employees"
    id = Column(Integer, primary_key=True)
    name = Column(String, unique=True)
    classification = Column(String)
    phone = Column(String)
    email = Column(String)
    active = Column(String, default="Yes")

class Equipment(Base):
    __tablename__ = "equipment"
    id = Column(Integer, primary_key=True)
    name = Column(String, unique=True)
    category = Column(String)
    active = Column(String, default="Yes")

class ScheduleItem(Base):
    __tablename__ = "schedule_items"
    id = Column(Integer, primary_key=True)
    schedule_date = Column(String)
    job_number = Column(String)
    resource_type = Column(String)
    resource_name = Column(String)
    is_lead = Column(String, default="No")

class AccountingJob(BaseModel):
    job_number: Optional[str] = ""
    job_name: Optional[str] = ""
    customer: Optional[str] = ""
    pm: Optional[str] = ""
    address: Optional[str] = ""
    city: Optional[str] = ""
    state_zip: Optional[str] = ""
    ret_notes: Optional[str] = ""

class ImportPayload(BaseModel):
    jobs: List[AccountingJob]

app = FastAPI()

@app.on_event("startup")
def startup():
    Base.metadata.create_all(bind=engine)

def clean(value):
    if value is None:
        return ""
    return str(value).strip()

def import_jobs_to_db(items):
    db = SessionLocal()
    imported = 0
    updated = 0
    skipped = 0

    try:
        for item in items:
            ret_notes = clean(item.get("ret_notes"))
            if ret_notes.lower() == "done":
                skipped += 1
                continue

            job_number = clean(item.get("job_number"))
            job_name = clean(item.get("job_name"))

            # Job Number is required for accounting imports.
            if not job_number:
                skipped += 1
                continue

            job = db.query(Job).filter(Job.job_number == job_number).first() if job_number else None

            if job:
                updated += 1
            else:
                job = Job(job_number=job_number)
                db.add(job)
                imported += 1

            job.job_name = job_name
            job.customer = clean(item.get("customer"))
            job.pm = clean(item.get("pm"))
            job.address = clean(item.get("address"))
            job.city = clean(item.get("city"))
            job.state_zip = clean(item.get("state_zip"))
            job.status = job.status or "Active"
            job.updated_at = datetime.utcnow()

        db.commit()
        return {"ok": True, "imported": imported, "updated": updated, "skipped": skipped}
    finally:
        db.close()

@app.get("/api/health")
def health():
    return {"ok": True, "app": "Marschel Operations Board"}

@app.get("/api/jobs")
def get_jobs():
    db = SessionLocal()
    try:
        rows = db.query(Job).order_by(Job.pm, Job.job_number).all()
        return {"ok": True, "jobs": [
            {
                "job_number": r.job_number,
                "job_name": r.job_name,
                "customer": r.customer,
                "pm": r.pm,
                "address": r.address,
                "city": r.city,
                "state_zip": r.state_zip,
                "start_date": r.start_date,
                "date_entered": r.date_entered,
                "duration_days": r.duration_days,
                "laborers_needed": r.laborers_needed,
                "operators_needed": r.operators_needed,
                "job_type": r.job_type,
                "priority": r.priority,
                "status": r.status,
            } for r in rows
        ]}
    finally:
        db.close()

@app.post("/api/import-accounting-jobs")
def import_accounting_jobs(payload: ImportPayload, x_import_token: str = Header(default="")):
    if IMPORT_TOKEN and x_import_token != IMPORT_TOKEN:
        raise HTTPException(status_code=401, detail="Invalid import token")
    return import_jobs_to_db([j.dict() for j in payload.jobs])

@app.post("/api/import-accounting-file")
async def import_accounting_file(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    content = await file.read()
    wb = load_workbook(BytesIO(content), data_only=True)

    ws = wb.active

    required_headers = [
        "Job #",
        "Job Name Description",
        "Customer",
        "Estimator/PM",
        "Address",
        "City/County",
        "State, Zip",
        "RET NOTES",
    ]

    header_row = None
    headers = {}

    for row in range(1, min(ws.max_row, 25) + 1):
        row_headers = {}
        for col in range(1, ws.max_column + 1):
            header = clean(ws.cell(row=row, column=col).value)
            if header:
                row_headers[header] = col

        if all(h in row_headers for h in required_headers):
            header_row = row
            headers = row_headers
            break

    if not header_row:
        raise HTTPException(status_code=400, detail="Could not find accounting header row with required columns")

    min_row = header_row
    max_row = ws.max_row

    required = [
        "Job #",
        "Job Name Description",
        "Customer",
        "Estimator/PM",
        "Address",
        "City/County",
        "State, Zip",
        "RET NOTES",
    ]

    missing = [h for h in required if h not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing headers: {', '.join(missing)}")

    items = []
    for row in range(min_row + 1, max_row + 1):
        items.append({
            "job_number": clean(ws.cell(row=row, column=headers["Job #"]).value),
            "job_name": clean(ws.cell(row=row, column=headers["Job Name Description"]).value),
            "customer": clean(ws.cell(row=row, column=headers["Customer"]).value),
            "pm": clean(ws.cell(row=row, column=headers["Estimator/PM"]).value),
            "address": clean(ws.cell(row=row, column=headers["Address"]).value),
            "city": clean(ws.cell(row=row, column=headers["City/County"]).value),
            "state_zip": clean(ws.cell(row=row, column=headers["State, Zip"]).value),
            "ret_notes": clean(ws.cell(row=row, column=headers["RET NOTES"]).value),
        })

    result = import_jobs_to_db(items)
    result["file"] = file.filename
    result["rows_read"] = len(items)
    return result

app.mount("/", StaticFiles(directory=".", html=True), name="static")


